import openpyxl as opx
import csv
import copy 
import pprint

# Converts csv to xlsx
def csv_to_excel(csv_file, excel_file):
    csv_data = []
    with open(csv_file) as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)
    workbook = opx.Workbook()
    sheet = workbook.active
    for row in csv_data:
        sheet.append(row)
    workbook.save(excel_file)

# print("Converting csv to xlsx")
# csv_to_excel("C:\\Users\\Guest User\\Desktop\\rangsee\\Accessories.csv", "C:\\Users\\Guest User\\Desktop\\rangsee\\Accessories.xlsx")
# print("Converted to xlsx")

# print("Opening Workbook")
# workbook = opx.load_workbook("C:\\Users\\Guest User\\Desktop\\rangsee\\Accessories.xlsx")
# print("Opened Workbook")
# ws = workbook.active
# handles = []

# print("Parsing doc")
# for row in ws.iter_rows(min_row = 2, max_col=1,max_row=9230, values_only=True):
#     for data in row:
#         if data not in handles:
#             handles.append(data)
# print("Done")            
# # print(f"Handles length: {len(handles)}")
# print(f"Handles length: {len(handles)}\n First item: {handles[0]}\n Last item {handles[-1]}")

# print("Creating new sheet")
# ws2 = workbook.create_sheet("Handles")
# print("Appending Handles to new sheet")

# col = 1
# for i, handle in enumerate(handles):
#     ws2.cell(row = i+1, column = col).value = handle 

# workbook.save("C:\\Users\\Guest User\\Desktop\\rangsee\\Accessoriess.xlsx")
# print("Done")

# print("Opening Workbook")
# workbook = opx.load_workbook("C:\\Users\\Guest User\\Desktop\\rangsee\\Samsungproducts_export_1.xlsx")
# print("Opened Workbook")
# ws = workbook.active
# handles = []

# reviews = opx.load_workbook("C:\\Users\\Guest User\\Desktop\\rangsee\\reviews.xlsx")
# ws = reviews.active

# galaxy = []
# for row in ws.iter_rows(min_row = 2, max_col = 7, values_only=True):
#     if ("Pill" in row[5].split(" ")) or ("Lipstick" in row[5].split(" ")) or ("Card" in row[5].split(" "))  :
#         galaxy.append(row)

# headers = []
# for col in ws.iter_cols(max_row = 1, max_col = 7, values_only=True):
#     headers.append(col)

formatted = opx.load_workbook("C:\\Users\\Guest User\\Desktop\\rangsee\\Accessoriess.xlsx")
handles = formatted["Handles"]
galaxy = []

for row in handles.iter_rows( max_col=1,max_row=256, values_only=True):   
    galaxy.append(row[0])
print(type(galaxy[0]))
    
print(f"Galaxy length: {len(galaxy)}\n First item: {galaxy[0]}\n Last item {galaxy[-1]}")

etsy_reviews = formatted["Reviews"]
reviews = []
for row in etsy_reviews.iter_rows(min_row = 2, max_col = 7, values_only=True):
    reviews.append(list(row))
print(len(reviews))	

all_data = []
print("values of galaxy")
for gal in galaxy:
    for rev in reviews:      
        new = copy.deepcopy(rev)
        del new[-1]
        new.append(gal)
        all_data.append(new)
print("Done values of galaxy")


# pprint.pprint(len(all_data))

# pprint.pprint(len(all_data[0]))
# for i, item in enumerate(all_data[0]):
#     print(i, item )

final = formatted["Formatted"]

print(len(all_data))
for i in range(2,len(all_data)):
    for val in all_data[i]:
        # print(f"un {all_data[i][0]}")
        title = final.cell(row = i, column = 1)
        title.value = all_data[i][5]
        body = final.cell(row = i, column =2)
        body.value = all_data[i][4]
        rating = final.cell(row = i, column =3)
        rating.value = all_data[i][3]
        date = final.cell(row=i, column = 4)
        date.value = all_data[i][2]
        rev_name = final.cell(row = i, column = 5)
        rev_name.value = all_data[i][0]
        rev_id = final.cell(row = i, column = 6)
        rev_id.value = all_data[i][1]
        handle = final.cell(row = i, column =8)
        handle.value = all_data[i][6]
#     i+=1
#     print(i)
    print(i)

# for i in len(all_data):
#     cellref = final.cell(row=i,)

print("Done!!!!!!!!!!!!!!!!!!!")
formatted.save("C:\\Users\\Guest User\\Desktop\\rangsee\\Accfinallll.xlsx")




